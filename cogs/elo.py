
"""
The MIT License (MIT)
Copyright (c) 2017 Dino
Permission is hereby granted, free of charge, to any person obtaining a
copy of this software and associated documentation files (the "Software"),
to deal in the Software without restriction, including without limitation
the rights to use, copy, modify, merge, publish, distribute, sublicense,
and/or sell copies of the Software, and to permit persons to whom the
Software is furnished to do so, subject to the following conditions:
The above copyright notice and this permission notice shall be included in
all copies or substantial portions of the Software.
THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS
OR IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING
FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER
DEALINGS IN THE SOFTWARE.
"""
from __main__ import send_cmd_help, settings
from bs4 import BeautifulSoup
from cogs.utils import checks
from cogs.utils.dataIO import dataIO
from copy import deepcopy
from discord.ext import commands
from openpyxl.utils import cell
from openpyxl import Workbook
import aiohttp
import asyncio
import discord
import inspect
import json
import locale
import math
import operator
import os
import requests
import time
import urllib
import urllib.request
import string

class GameException(Exception):
	def __init__(self, msg):
		self.msg = msg

def timit(func, *args, **kwargs):
	def wrapper(args, **kwargs):
		t1 = time.time()
		retval = func(args, **kwargs)
		t2 = time.time()
		print(func.__name__,':',round(t2-t1,6), 'seconds')
		return retval
	return wrapper

# name_of_num = {
# 	0:'zero',
# 	1:'one',
# 	2:'two',
# 	3:'three',
# 	4:'four',
# 	5:'five',
# 	6:'six',
# 	7:'seven',
# 	8:'eight',
# 	9:'nine',
# }
lb_channel = '391457160783069184'
submit_channels = ['388767658214162433', '390015530150658048']
whitelisted = [
	'372659236683120670',
	'388767658214162433',
	'3885762542086389760', 
	'264119826069454849',
	'388576254208638976',
	'389786665171943425',
	'389281078684286978', #bot
	'305591335111360512',
	'390043766498656266', #dino
	'388767658214162433', #submit-results
	#whitelisted channels
]
def is_owner_or_Piripic_check(ctx):
    print(ctx.message.author.id in [settings.owner, '235334114390048769'])
    print(ctx.message.author.id)
    print(settings.owner, '235334114390048769')
    return ctx.message.author.id in [settings.owner, '235334114390048769']

PATH = os.path.join("data", "league")
DATA_JSON = os.path.join(PATH, "data.json")
LOGS_JSON = os.path.join(PATH, "logs.json") #not in use
TEMP_JSON = os.path.join(PATH, "temp.json")
WHEN_JSON = os.path.join(PATH, "when.json")
ALLSTAT_EXCEL = os.path.join(PATH, "elo.xlsx")
when_check_interval = 16 * 60 #16 minutes, in seconds
when_interval = 15 * 60 #15 minutes, in seconds
default_when_json = {'time': time.time()}
default_data_json = {'server':{'allnames':[], 'idtoname':{}, 'games':[], 'trophies':{}, 'cooldowns':{}, 'logs':[], 'queue':[], 'denied':[]}}
#list of all names registered
#dict of discord id to discord name (only ascii)
#list of 3 length lists showing ['winner', 'loser', 'inputter']
#dict of all players and their trophies
games_in_time = 1
timeframe = 1*3*60*60 # 3 hrs in seconds
default_logs_json = {'server':[]}
#list of dicts in format:
#{'code': 'a|ac|cb|e', 'logger': 'discordid', 'winner': 'playername', 'loser': 'playername'}
log_codes = {
	'a': 'attempted',
	'ae': 'attempted with an error',
	'c': 'attempted during cooldown',
	'bc': 'cooldown bypassed (mods/admins)',
	'e': 'executed',
	'add': 'admin hard-added trophies to a person',
	'd': 'denied by non-logger',
}
#starts with logcode means log_codes[logcode]

COOLDOWN_JSON = os.path.join(PATH, "cooldown.json")
cooldowntime = 3 * 60 #3 minutes (in seconds)
queue_responsetime = 4 * 60 * 60 #4 hours (in seconds)
default_cooldown_json = {'server':{}}
#'discordid': time since 1970

trophies_per_win = 30 #like CR ladder or tournaments 
def diff_to_change(diff):
	return diff/10
#if this value func is x: x/10 40 trophy person vs 30 trophy person:
#40 person wins: gets 29; 30 person loses 29
#30 person wins: gets 31; 40 person loses 31
niceascii = string.ascii_uppercase + string.ascii_lowercase + string.digits + string.punctuation
def remove_non_ascii(thing): 
	return str(list(filter(lambda x: x in niceascii, thing)))

class League:

	def queuegame(self,ctx, p2id, state):
		author = ctx.message.author
		server = ctx.message.server
		self.settings[server.id]['queue'].append([author.id,p2id,state,time.time()])
		self.save_settings()

	async def check_queue(self):
		# author = ctx.message.author
		# server = ctx.message.server
		for server in self.bot.servers:
			n = 0
			for queuedgm in self.settings[server.id]['queue']:
				# print(queuedgm, queuedgm[3], type(queuedgm[3]))
				if time.time() - queuedgm[3] > queue_responsetime:


					thing = self.settings[server.id]['queue'].pop(n)
					if thing[2] == 'w':
						winner, loser = thing[0], thing[1]
					else:
						winner, loser = thing[1], thing[0]
					# self.settings[server.id]['queue'].pop(listnum)
					await self.bot.send_message(server.get_member(thing[0]), '{} has denied(because of queue response time timeout) your {} against them. If this game is legitiment(shouldn\'t have been denied), please DM someone with the `ELO Support` role with screenshots of the bo3.'.format(self.settings[server.id]['idtoname'][thing[1]], 'win' if thing[2]=='w' else 'loss'))
					self.settings[server.id]['denied'].append({'winner':winner, 'loser':loser,'logger':thing[0], 'code':'d'})
					self.save_settings() 
				n += 1


	def load_settings(self):
		self.settings = dataIO.load_json(DATA_JSON)

	def save_settings(self):
		# self.settings
		# print(self.settings)
		try:
			dataIO.save_json(DATA_JSON, self.settings)
		except TypeError:
			print(self.settings)
		# print(self.settings == dataIO.load_json(DATA_JSON))

	def log_to_dict(self, ctx, loglist):
		return ctx, {'code': loglist[0], 'logger': loglist[1], 'winner': loglist[2], 'loser': loglist[3],}

	def log(self, ctx, logdict):
		if type(logdict) == type([]): #if it is a list
			ctx, logdict = self.log_to_dict(ctx, logdict)
		self.settings[ctx.message.server.id]['logs'].append(logdict)
		self.save_settings()

	def add_cooldown(self, ctx):
		author = ctx.message.author
		server = ctx.message.server
		self.settings[server.id]['cooldowns'][author.id] = time.time()
		self.save_settings()

	def populate_settings(self):
		if len(list(self.bot.servers))>0:
			for server in self.bot.servers:
				if server.id not in self.settings or self.settings[server.id] == {}:
					self.settings[server.id] = deepcopy(default_data_json['server'])#{'allnames':[], 'idtoname':{}, 'games':[], 'trophies':{}, 'cooldowns':{}, 'logs':[] }
				# self.settings[server.id] = default_data_json['server'].copy()
		for server in self.bot.servers:
			sid = server.id
			for key in default_data_json['server']:
				if key not in self.settings[sid]:
					self.settings[sid][key] = deepcopy(default_data_json['server'][key])
		# print(self.bot)
		# print(self.bot.servers)
		# print('penis', self.bot.servers)
		self.save_settings()

	def populate_server_settings(self, serverid):
		if self.settings[serverid] == {}:
			self.settings[serverid] = deepcopy(default_data_json['server'])#{'allnames':[], 'idtoname':{}, 'games':[], 'trophies':{}, 'cooldowns':{},  'logs':[]}
		self.save_settings()

	def __init__(self, bot):
		#note self.settings hierarchy below
		# for server in self.settings:
		# 	self.settings[server.id]['idtoname'] = self.settings[server.id]['idtoname']
		# 	self.settings[server.id]['games'] = self.settings[ctx.message.server.id]['games']
		# 	self.settings[server.id]['allnames'] = self.settings[server.id]['allnames']
		# 	self.settings[server.id]['trophies'] = self.settings[server.id]['trophies']
		# time.sleep(5) #need bot to load fully

		# def change(func, *args, **kwargs):
		# 	def wrapper(args, **kwargs):
		# 		# tmp = {
		# 		# 	'bot': self,
		# 		# 	'invoked_with': invoker,
		# 		# 	'message': message,
		# 		# 	'view': view,
		# 		# 	'prefix': invoked_prefix
		# 		# }
		# 		# tmpctx = Context(**tmp)
		# 		# del tmp
		# 		self.populate_settings()
		# 		yield from bot.get_command("elo registerserver").callback()

		# 		func(args, kwargs)
		# 	return wrapper
		# bot.on_message = change(bot.on_message)

		self.do_loop = {}
		self.bot = bot
		self.load_settings()
		def my_decorator(some_function):

			async def wrapper(*args, **kwargs):

				await some_function(*args, **kwargs)
				self.populate_settings()
				datachannel = bot.get_channel('389786665171943425')
				async for msg in self.bot.logs_from(datachannel):
					if len(msg.attachments)>0:
						break
				try:
					await self.recover(msg)
					print('recovered')
				except Exception as e:
					print(e)
				for server in self.settings:
					if 'games' in self.settings[server]:
						n = 0
						for game in self.settings[server]['games']:
							if len(game) == 2:
								self.settings[server]['games'][n].append(0)
							n += 1
				for server in bot.servers:
					if server.id in self.settings:
						self.settings[server.id]['allnames'] = []
						for mem in server.members:
							if 'idtoname' not in self.settings[server.id]:
								self.settings[server.id]['idtoname'] = {}
							self.settings[server.id]['idtoname'][mem.id] = mem.name
							self.settings[server.id]['allnames'].append(mem.name)
				self.save_settings()
				# print('throat pees', self.settings['trophies'])
				# self.populate_settings()
				# for server in bot.servers:
				# 	self.registerservermems(server)
				async def on_member_join(self, member):
					server = member.server
					self.registermem(server, member.id, member.name)
					self.save_settings()
				bot.on_member_join = on_member_join
				# print('herro')
				lb = bot.get_channel(lb_channel)
				# await self.bot.send_message(datachannel, 'penis')
				self.do_loop[datachannel.id] = True
				print('wedone')
				while self.do_loop[datachannel.id]:
				# 	# print('herro2')
				# 	await self.check_queue()
				# 	# await self.bot.send_message(datachannel, 'penis2')

				# 	if lb != None:
				# 		await self.refresh_leaderboard(lb)
				# 	# await self.bot.send_message(datachannel, 'penis3')
					await self.try_send_json(datachannel)
				# 	# await self.bot.send_message(datachannel, 'penis4')
				# 	for server in bot.servers:
				# 		self.registerservermems(server)
				# 		self.rerungames(server)
				# 	# await self.bot.send_message(datachannel, 'penis5')
					await asyncio.sleep(when_check_interval)
			return wrapper
			#loads some CRTags atttributes when bot is ready, otherwise it cant get servers/emojis
		self.bot.on_ready = my_decorator(self.bot.on_ready) #adds the above code to on_ready

# 		def my_decorator2(some_function):

# 			async def wrapper(*args, **kwargs):
# 				msg = args[0]
# 				chan = msg.channel
# 				if msg.channel.name == None:
# 					return
# 				# print(args[0])
# 				# print(dir(args[0]))

# 				# print(chan.name)
# 				# print('elo')
# 				# print('elo' not in chan.name)

# 				if chan.id not in whitelisted and 'elo' not in chan.name:
# 					return
# 				await some_function(*args, **kwargs)

# 			return wrapper
# 			#loads some CRTags atttributes when bot is ready, otherwise it cant get servers/emojis
# 		self.bot.on_message = my_decorator2(self.bot.on_message) #adds the above code to on_message

# 		# def my_decorator3(some_function):

# 		# 	async def wrapper(*args, **kwargs):
# 		# 		member = args[0]
# 		# 		server = member.server
# 		# 		self.registermem(server, member.id, member.name)
		#		self.save_settings()
# 		# 		await some_function(*args, **kwargs)

# 		# 	return wrapper
# 		# 	#loads some CRTags atttributes when bot is ready, otherwise it cant get servers/emojis
# 		# self.bot.on_member_join = my_decorator3(self.bot.on_member_join) #adds the above code to on_message
# 		# def my_decorator2(some_function):

# 		# 	async def wrapper(*args, **kwargs):
# 		# 		if args[0].author.id == '222925389641547776':
# 		# 			print(args[0].channel.id)
# 		# 		# if args[0].channel.id not in whitelisted:
# 		# 		# 	return
# 		# 		await some_function(*args, **kwargs)

# 		# 	return wrapper
# 			#loads some CRTags atttributes when bot is ready, otherwise it cant get servers/emojis
# 		# self.bot.on_message = my_decorator2(self.bot.on_message) #adds the above code to on_message

# 		self.pre = bot.settings.prefixes[0]
# 		# with open(DATA_JSON, 'r') as f:
# 			# print(f.read())
# 		# print(self.settings)
# 		self.populate_settings()
# 		# print(self.settings)
# 		# with open(DATA_JSON, 'r') as f:
# 		# 	print(f.read())
# 		funcs = dir(self)
# 		n=0
# 		for f in funcs:
# 			if '__' in f:
# 				funcs.pop(n)
# 			n+=1
# 		f = lambda x:x
# 		f = type(f)
# 		print('init finished')
# 		# for func in funcs:
# 		# 	if '{0}' in func.__doc__:
# 		# 		func.__doc__ = func.__doc__.format(self.pre)
# 		# for func in funcs:
# 		# 	if '' in func.__doc__:
# 		# 		func.__doc__ = func.__doc__.format(self.pre)
# 		# self.settings[server.id]['logs'] = dataIO.load_json(LOGS_JSON)
# 		# self.settings[server.id]['cooldowns'] = dataIO.load_json(COOLDOWN_JSON)

	@commands.command()
	async def testfunc(self):
		await self.bot.say('test')

	def registermem(self, server, discordid, name):
		# print('brls', server.id, discordid)
		# print(self.settings)
		# print(self.settings)
		if name not in self.settings[server.id]['allnames']:
			self.settings[server.id]['allnames'].append(name)
		self.settings[server.id]['idtoname'][discordid] = name

			# print('adding', 'name')
		if discordid not in self.settings[server.id]['allnames']:
			self.settings[server.id]['idtoname'][discordid] = name
		if discordid not in self.settings[server.id]['allnames']:
			self.settings[server.id]['trophies'][discordid] = 0
			# self.save_settings()

	def checkplayers2(self, ctx, players):
		server = ctx.message.server
		n = 0
		for p in players:
			if type(p) == type(discord.Member) or type(p) == type(discord.User):
				players[n] = p.name
			if p.startswith('<@') and p.endswith('>'): #mention
				p = p[2:-1]
				if p.startswith('!'): #mention with '!'
					p = p[1:]
				players[n] = p
			n+=1
		n = 0
		for p in players:
			if p in self.settings[server.id]['idtoname']:
				players[n] = self.settings[server.id]['idtoname'][p]
			elif p not in self.settings[server.id]['allnames']:
				raise GameException('{} is not a registered name or id.'.format(p))
				return
			n+=1
		return players

	def checkplayers(self, ctx, players):
		server = ctx.message.server
		n = 0
		for p in players:
			if 'discord.member' in str(type(p)) or 'discord.user' in str(type(p)):
				players[n] = p.id
			elif p.startswith('<@') and p.endswith('>'): #mention
				p = p[2:-1]
				if p.startswith('!'): #mention with '!'
					p = p[1:]
				players[n] = p
			n+=1
		n = 0
		for p in players:
			if p not in self.settings[server.id]['idtoname']:
				raise GameException('{} is not a registered  id.'.format(p))
				return
			n+=1
		return players

# 	@checks.admin()
# 	@commands.command(pass_context=True)
# 	async def recoverjson(self, ctx):
# 		server = ctx.message.server
# 		author = ctx.message.author
# 		if author.id not in ['222925389641547776', '235334114390048769']:
# 			return

# 		if await self.confirmation(ctx) == None:
# 			return
# 		try:
# 			await self.recover(ctx.message)
# 		except GameException as e:
# 			await self.bot.say(e)

	async def recover(self, msg):
		print('asdadc')
		filepath = msg.attachments[0]['url']
		if not filepath.endswith('.json'):
			# print('herro5')
			raise GameException('File must be json.')
		async with aiohttp.ClientSession() as session:
			async with session.get(filepath) as resp:
				# print('herro4')
				data = await resp.text()
				# data = str(data)[2:-1]
				# print(data[:100])
				# print('herro5')
		# print('herro6')
		data = json.loads(data)
		# print('herro7')
		# print('herro3', data['trophies'])
		self.settings = data
		# print('herro8')
		# print('servers:', ', '.join([server.name for server in self.bot.servers]))
		for server in self.bot.servers:
			# print('naem', server.name)
			self.rerungames(server)

		# print('trophies:', self.settings['trophies'])
		self.save_settings()


# 	@checks.is_owner()
# 	@commands.command(pass_context=True)
# 	async def test(self, ctx):
# 		"""test doc"""
# 		server = ctx.message.server
# 		author = ctx.message.author.bot.get_message(ctx.message.channel, '389119480518672385')
# 		# print(dir(msg))
# 		# print(msg.attachments)
# 		# print(msg.embeds)
# 		# print('i am a penis')
# 		# print(inspect.getsource(self.bot.on_message))
# 		# print('pre',self.pre)
# 		# print('test',self.test.__doc__)
# 		# print('default',self.defaultfunc.__doc__)
# 		# print('win',self.win.__doc__)
# 		# print('rep',self.win.__doc__.replace(self.defaultfunc.__doc__,''))

# 	@checks.admin()
# 	@commands.command(aliases=['sendserverjson'], pass_context=True)
# 	async def send_server_json(self, ctx):
# 		"""sends json for current server."""
# 		server = ctx.message.server
# 		author = ctx.message.author
# 		tempjson = {}
# 		tempjson[server.id] = self.settings[server.id]
# 		dataIO.save_json(TEMP_JSON, tempjson)
# 		await self.bot.send_file(author, TEMP_JSON)

	@checks.admin()
	@commands.command(aliases=['sendfulljson'], pass_context=True)
	async def send_full_json(self, ctx):
		server = ctx.message.server
		author = ctx.message.author
		"""sends full json."""
		# if not is_owner_or_Piripic_check(ctx):
		# 	return
		await self.bot.send_file(author, DATA_JSON)

	@commands.group(aliases=["elo"], pass_context=True)
	async def ELO(self, ctx):
		"""ELO League Management
		"""
		# await self.bot.delete_message(ctx.message)

		# subname = ctx.invoked_subcommand.name
		if ctx.invoked_subcommand is None:
			helptxts = [
				'Do `!elo playerhelp` if you are a regular player',
				'Do `!elo adminhelp` if you are a admin',
				'Do `!help elo` for cmds available to you.'
			]
			await self.bot.say(' or\n'.join(helptxts))
		# elif 'help' in subname:
		# 	await send_cmd_help(ctx)
		# print(subname)

	
# 	# @ELO.command(name='help',pass_context=True)
# 	# async def elo_help(self, ctx):
# 	# 	pass

	@checks.admin()
	@ELO.command(pass_context=True)
	async def adminhelp(self, ctx):
		"""help for admin(manager) commands.
		"""
		commands = {
			self.pre+'elo registerserver' : 'populates server with necessary data to add games, show lb, etc',
			self.pre+'elo game' : 'adds game of winner vs loser',
			self.pre+'elo add' : 'adds (can subtract with negative) trophies to certain person. This add is ignored if rereungames is run.',
			self.pre+'elo rerungames' : 'reruns all games, useful after removing a game',
			self.pre+'elo rmlog' : 'removes log at index. Useful with !elo showlogs',
			self.pre+'elo rmgame' : 'removes game at index. Useful with !elo showgames',
			self.pre+'elo rmdeny' : 'removes denied game at index. Useful with !elo showdenied',
			self.pre+'elo undeny' : 'makes denied game at index to a regular game. Useful with !elo showdenied',
			self.pre+'elo showlogs' : 'sends a list of all actions logged',
			self.pre+'elo showgames' : 'sends a list of all games played',
			self.pre+'elo showusergames' : 'sends a list of all games played by p1',
			self.pre+'elo showdenied' : 'sends a list of all games denied',
			self.pre+'elo showqueue' : 'sends a list of all games queued',
			self.pre+'elo rmqueue' : 'removes queued game at that index, useful with !elo showgames',
			self.pre+'elo confirmgame' : 'confirms game from p1\'s perspective that p2 requested.Useful with !elo showqueue',
			self.pre+'elo resetgames' : 'resets all games for current server',
			self.pre+'elo resetall' : 'resets all data for current server',
			self.pre+'elo savejson' : 'saves the current json.',
			self.pre+'autosendjson' : 'auto-sends a json of all the players every {} seconds'.format(when_interval),
			self.pre+'stopautosendjson' : 'un-does the action of the above cmd.',
			self.pre+'sendfulljson' : 'sends full json of ELO league',
			self.pre+'pop' : 'populates json with the starting dicts',
			self.pre+'adminhelp' : 'sends this command',
		}
		longest = ''
		n=0
		for cmd in commands:
			if len(cmd)>len(longest):
				longest = cmd
			n+=1
		maxlen = len(longest)+2
		beg = '```\n'
		end = '```'
		helptxt = ''
		helptxt += 'ELO admin commands:\n\n'
		for cmd in commands:
			spaces = maxlen-len(cmd)
			helptxt += "  {}{}{}".format(cmd,spaces*' ', commands[cmd])
			helptxt += '\n'
		# helptxts = helptxt.split('\n')
		# helptxts2 = []
		# helptxts = [helptxts[:int(len(helptxts)/2)],helptxts[int(len(helptxts)/2):]]
		helptxt = beg+helptxt+end
		# msgsreq = math.ceil(len(helptxt/2000))
		# x = 7
		# n = 0
		# while x<2000:
		# 	x+=len(helptxts[n])+1
		# 	n+=1
		await self.bot.say(helptxt)

		# with open(TEMP_JSON, 'w+') as f:
		# 	f.write(helptxt)
		# await self.bot.send_file(ctx.message.author, TEMP_JSON)
		# for txt in helptxts:
		# 	htxt = ''
		# 	htxt += '```\n'
		# 	htxt = '\n'.join(txt)
		# 	htxt += '```'
		# 	await self.bot.send_message(ctx.message.author, htxt)

	@ELO.command(pass_context=True)
	async def playerhelp(self, ctx):
		"""help for players in the league
		"""
		commands = {
			self.pre+'elo rules' : 'shows rules of ELO League',
			self.pre+'elo lose' : 'queues up a loss for yourself vs p2, requires other player to confirm that this game has been played',
			self.pre+'elo win' : 'queues up a win for yourself vs p2, requires other player to confirm that this game has been played',
			self.pre+'elo confirm' : 'confirms someone else\'s queued game vs you',
			self.pre+'elo deny' : 'denies someone else\'s queued game vs you',
			self.pre+'elo rank' : 'Shows your rank among all players in the League',
			self.pre+'elo lb' : 'Shows leaderboard',
			self.pre+'elo mygames' : 'Shows all the games that you have played',
			self.pre+'elo playerhelp' : 'shows this message',
			self.pre+'elo register' : 'registers yourself. you probably wont need this, unless you just joined',
		}
		longest = ''
		n=0
		for cmd in commands:
			if len(cmd)>len(longest):
				longest = cmd
			n+=1
		maxlen = len(longest)+2
		beg = '```\n'
		end = '```'
		helptxt = ''
		helptxt += 'ELO player commands:\n\n'
		for cmd in commands:
			spaces = maxlen-len(cmd)
			helptxt += "  {}{}{}".format(cmd,spaces*' ', commands[cmd])
			helptxt += '\n'
		# helptxts = helptxt.split('\n')
		# helptxts2 = []
		# helptxts = [helptxts[:int(len(helptxts)/2)],helptxts[int(len(helptxts)/2):]]
		helptxt = beg+helptxt+end
		# msgsreq = math.ceil(len(helptxt/2000))
		# x = 7
		# n = 0
		# while x<2000:
		# 	x+=len(helptxts[n])+1
		# 	n+=1
		await self.bot.say(helptxt)
		# with open(TEMP_JSON, 'w+') as f:
		# 	f.write(helptxt)
		# await self.bot.send_file(ctx.message.author, TEMP_JSON)
		# for txt in helptxts:
		# 	htxt = ''
		# 	htxt += '```\n'
		# 	htxt = '\n'.join(txt)
		# 	htxt += '```'
		# 	await self.bot.send_message(ctx.message.author, htxt)

	@ELO.command(pass_context=True)
	async def rules(self,ctx, dm=None):
		""" DM's rules
		"""

		server = ctx.message.server
		author = ctx.message.author
		mod_role = settings.get_server_mod(server).lower()
		admin_role = settings.get_server_admin(server).lower()
		if dm != None:
			roles = list(map(lambda role: role.name.lower(), author.roles))
			if not(admin_role in roles or checks.is_owner_check(ctx)):
				dm = None

		rules = [
			'Co-ordinate with your opponent after the match to determine who enters the match so it doesn\'t get double entered.',
			'Cooldown is 1 hour, if you play a match again within an hour, ask a Manager to enter the match.',
			'This ELO system will provide a dynamic, weighted power-ranking score for all those participating, which we want to be all players. Simply play another player on this server in a Best of 3 games and make sure to tag your opponent in #score_submissions along with who won to make sure that they agree that they intended to play an ELO match with you and it was not supposed to be a practice game. If there is a dispute screenshots may be asked for by a @Manager.',
			'After entering a match, both players should react with a checkmark on the #score_submissions message.'
			'The top 3 performers each month will get a prize reward of $50 , $30 , $20 respectively',
			'Every player who has signed up must play a minimum of 3 matches a week. You can only play the same opponent a maximum of twice a week. ',
			'After each match in the best of 3, the loser of the match may choose to ban a card which will be banned for the remainder of the best of 3. ',
			'If you are caught to be cheating or abusing the system/bot you will be banned from earning prizes from the ELO system and will not be permitted to play in it until a @Manager feels you have been suspended for long enough. The minimum suspension is 1 month and the maximum is a permanent ban from the ELO system.',
			'We will begin this ELO system on Monday the 11th of December and will end after 31 days',
		]
		rulestxt = ''
		rulestxt += '```py\n'
		n = 0
		for rule in rules:
			rulestxt += "———\n[{}]— {}\n".format(n+1,rule)
			n += 1
		rulestxt += '```'
		if dm == None:
			await self.bot.send_message(ctx.message.author, rulestxt)
		else:
			await self.bot.say(rulestxt)

# 	# @ELO.command(aliases=['commands'], pass_context=True)
# 	# async def cmds(self, ctx):
# 	# 	server = ctx.message.server
# 	# 	author = ctx.message.author
# 	# 	mod_role = settings.get_server_mod(server).lower()
# 	# 	admin_role = settings.get_server_admin(server).lower()
# 	# 	allcmds = self.bot.commands
# 	# 	# print(__file__)
# 	# 	# file = __file__
# 	# 	# while '\\' in file:
# 	# 	# 	file = file[file.find('.')+1:]
# 	# 	# print('lol')
# 	# 	curcmds = []
# 	# 	cla = str(self.__class__)
# 	# 	cla = cla.replace('<class \'','').replace('\'>','')
# 	# 	while '.' in cla:
# 	# 		cla = cla[cla.find('.')+1:]
# 	# 	for cmd in allcmds:
# 	# 		if cmd in dir(self):
# 	# 			curcmds.append(cmd)
# 	# 	msg = '```\nAll commands in {}:\n'.format(cla)
# 	# 	msg += ', '.join(curcmds)
# 	# 	msg += '```'
# 	# 	# cog = self.bot.cogs[cla]
# 	# 	await self.bot.say(msg)
	def registerservermems(self, server):
		for member in server.members:
			if member.bot:
				continue
			name = member.name
			self.registermem(server, member.id, name)
		self.save_settings()

	@checks.admin()
	@ELO.command(pass_context=True)
	async def registerserver(self, ctx):
		"""registers all of current server's members in the database
		"""
		server = ctx.message.server
		author = ctx.message.author
		mod_role = settings.get_server_mod(server).lower()
		admin_role = settings.get_server_admin(server).lower()
		self.registerservermems(server)
		await self.bot.say('Registered this server, `{}`.'.format(server.name))

	@ELO.command(pass_context=True)
	async def register(self, ctx):
		"""register yourself in the database
		"""
		server = ctx.message.server
		author = ctx.message.author
		mod_role = settings.get_server_mod(server).lower()
		admin_role = settings.get_server_admin(server).lower()
		if author.id in self.settings[server.id]['allnames']:
			await self.bot.say('Already Registered.')
			return
		self.registermem(server, author.id, author.name)
		self.save_settings()
		await self.bot.say('Registered.')

	def add_game(self,server, winner:str, loser, tim=None, position:int=None):
		# server = ctx.message.server
		if tim == None:
			tim = time.time()
		gam = [winner, loser, tim]
		if position == None:
			self.settings[server.id]['games'].append(gam)#time.time()])
		else:
			self.settings[server.id]['games'].insert(position, gam)#time.time()])
		if type(loser) == type(''):
			diff = self.settings[server.id]['trophies'][winner] - self.settings[server.id]['trophies'][loser]
			change = diff_to_change(diff)
			trophieschanged = round(trophies_per_win-change)
			#final trophy addition for winner
			minwin = 10
			maxwin = 50
			if trophieschanged<minwin:
				trophieschanged = minwin
			elif trophieschanged>maxwin:
				trophieschanged = maxwin
			#max you can win/lose is 50, min is 10


			self.settings[server.id]['trophies'][winner] += trophieschanged	#gain trophies
			self.settings[server.id]['trophies'][loser] -= trophieschanged		#lose trophies
			if self.settings[server.id]['trophies'][loser] <0:
				self.settings[server.id]['trophies'][loser] = 0
		else:
			self.settings[server.id]['trophies'][winner] += loser
		# print('adding trophies to {} and removing from {}'.format(winner, loser))
		# self.save_settings()
		# self.rerungames(server)
		return gam

	@checks.mod()
	@ELO.command(pass_context=True) #todo add cooldown 300
	async def game(self, ctx, p1:discord.Member, p2:discord.Member):
		"""registers a game of p1 winning vs p2
		"""
		server = ctx.message.server
		author = ctx.message.author
		mod_role = settings.get_server_mod(server).lower()
		admin_role = settings.get_server_admin(server).lower()
		players = [p1.id,p2.id]
		try:
			players = self.checkplayers(ctx, players)
		except GameException as e:
			await self.bot.say(e)
			self.log(ctx, ['ae', author.id, p1.id, p2.id])
			return
		winner, loser = players
		# if author.id in self.settings[server.id]['cooldowns']:
		# 	cd = time.time() - self.settings[server.id]['cooldowns'][author.id]
		# 	m, s = divmod(seconds, 60)
		# 	h, m = divmod(m, 60)
		# 	if cd > cooldowntime:
		# 		roles = list(map(lambda role: role.name, author.roles))
		# 		if mod_role in roles or admin_role in roles:
		# 			self.log(ctx, ['bc', author.id, winner, loser])
		# 		else:
		# 			await self.bot.say("Cooldown of 1 hour still in effect, {}:{}:{} left. If you need to enter a score, contact a Manager".format(h,m,s))
		# 			self.log(ctx, ['c', author.id, winner, loser])
		# 			return

		self.add_game(server, winner, loser)
		await self.bot.say('Game has been added: <@{}> beat <@{}>'.format(winner,loser))
		self.log(ctx, ['e', author.id, winner, loser])

	def double_reg(self, server, p1, p2):
		idtoname = self.settings[server.id]['idtoname']
		allqueue = self.settings[server.id]['queue']
		cur_time = time.time()
		for game in allqueue:
			# print(game)
			if p1 in game and p2 in game:
				# print('mygame', p1, p2)
				# print(game[3]-  cur_time, timeframe, game[3]-cur_time<timeframe)
				# if cur_time-game[3]<timeframe:
				return [True, '{} has already registered a {} against {}, would you like to queue up another game with them?'.format(idtoname[game[0]], 'win' if game[2]=='w' else 'loss', idtoname[game[1]])]
		# if nqueue>= games_in_time:
			# return False
		return [False]

	def game_allowed(self,server,p1,p2):
		allgames = self.settings[server.id]['games']
		allqueue = self.settings[server.id]['queue']
		cur_time = time.time()
		ngames = 0
		# print('p1',p1,'p2',p2)
		for game in allgames:
			if p1 in game and p2 in game and cur_time-game[2]<timeframe:
				ngames += 1
		if ngames>= games_in_time:
			return False
		nqueue = 0
		# print(allqueue)
		for game in allqueue:
			# print(game)
			if p1 in game and p2 in game:
				# print('mygame', p1, p2)
				# print(game[3]-  cur_time, timeframe, game[3]-cur_time<timeframe)
				if cur_time-game[3]<timeframe:
					nqueue += 1
		if nqueue>= games_in_time:
			return False
		return True


	@ELO.command(pass_context=True) #todo add cooldown 300
	async def win(self, ctx, p2:discord.Member):
		"""registers a win of yourself vs p2
		"""
		if ctx.message.channel.id not in submit_channels:
			return
		server = ctx.message.server
		author = ctx.message.author
		mod_role = settings.get_server_mod(server).lower()
		admin_role = settings.get_server_admin(server).lower()
		p1 = author
		players = [p1,p2]
		# print(type(p2))
		try:
			players = self.checkplayers(ctx, players)
		except GameException as e:
			await self.bot.say(e)
			self.log(ctx, ['ae', author.id, p1, p2])
			return
		# print(author.id)
		# print(self.settings[server.id]['cooldowns'])
		# print(author.id in self.settings[server.id]['cooldowns'])
		# print(checks.is_owner_check(ctx))
		# games in an amount of time vs same person
		thing = self.double_reg(server, p1.id, p2.id)
		if thing[0]:
			# await self.bot.say(thing[1])
			if await self.confirmation(ctx, confirmmsg=thing[1]) == None:
				return
			
		# print('HERRO')
		if not self.game_allowed(server, p1.id, p2.id):
			# print('not allowed')
			cd = timeframe
			m, s = divmod(cd, 60)
			h, m = divmod(m, 60)
			await self.bot.say("You may not play more than {} games in {:02d}:{:02d}:{:02d} time with the same person".format(games_in_time, h,m,s))
			return
		# else:
		# 	print('allowed')
		winner, loser = players
		if author.id in self.settings[server.id]['cooldowns']:
			cd = time.time() - self.settings[server.id]['cooldowns'][author.id]
			timeleft = cooldowntime-cd
			m, s = divmod(timeleft, 60)
			h, m = divmod(m, 60)
			# print('difference in time:{}:{}:{}'.format(h,m,s))
			if timeleft>0:
				roles = list(map(lambda role: role.name.lower(), author.roles))
				if mod_role in roles or admin_role in roles or checks.is_owner_check(ctx):
					# print('-------------\nlogging cypass cooldown for loss\n------------')
					self.log(ctx, ['bc', author.id, winner, loser])
				else:
					await self.bot.say("Cooldown of 1 hour still in effect, {:02d}:{:02d}:{:02d} left. If you need to enter a score, contact a Manager".format(int(h),int(m),round(s)))
					self.log(ctx, ['c', author.id, winner, loser])
					return

		self.queuegame(ctx, loser, 'w')
		await self.bot.say('Game has been queued, awaiting confirmation: <@{}> beat <@{}>'.format(winner,loser))
		other = server.get_member(loser)
		state = 'win'
		await self.bot.send_message(other, '<@{}> registered a {} against you. type `!elo confirm `<@{}>` ` to confirm this.'.format(author.id,state,author.id))
		self.add_cooldown(ctx)

	@ELO.command(pass_context=True) #todo add cooldown 300
	async def lose(self, ctx, p2:discord.Member):
		"""registers a loss of yourself vs p2
		"""
		
		if ctx.message.channel.id not in submit_channels:
			return
		# p2 = p2
		server = ctx.message.server
		author = ctx.message.author
		mod_role = settings.get_server_mod(server).lower()
		admin_role = settings.get_server_admin(server).lower()
		p1 = author
		players = [p2,p1]
		try:
			players = self.checkplayers(ctx, players)
		except GameException as e:
			await self.bot.say(e)
			self.log(ctx, ['ae', author.name, p1, p2])
			return
		# print(author.id)
		# print(self.settings[server.id]['cooldowns'])
		# print(author.id in self.settings[server.id]['cooldowns'])
		# print(checks.is_owner_check(ctx))
		# games in an amount of time vs same person
		thing = self.double_reg(server, p1.id, p2.id)
		if thing[0]:
			await self.bot.say(thing[1])
			if await self.confirmation(ctx) == None:
				return

		if not self.game_allowed(server, p1.id, p2.id):
			cd = timeframe
			m, s = divmod(cd, 60)
			h, m = divmod(m, 60)
			await self.bot.say("You may not play more than {} games in {:02d}:{:02d}:{:02d} time with the same person".format(games_in_time, h,m,s))
			return
		winner, loser = players
		if author.id in self.settings[server.id]['cooldowns']:
			cd = time.time() - self.settings[server.id]['cooldowns'][author.id]
			m, s = divmod(cd, 60)
			h, m = divmod(m, 60)
			# print('difference in time:{}:{}:{}'.format(h,m,s))
			if cd < cooldowntime:
				roles = list(map(lambda role: role.name.lower(), author.roles))
				if mod_role in roles or admin_role in roles or checks.is_owner_check(ctx):
					# print('-------------\nlogging cypass cooldown for loss\n------------')
					self.log(ctx, ['bc', author.id, winner, loser])
				else:
					await self.bot.say("Cooldown of 1 hour still in effect, {:02d}:{:02d}:{:02d} left. If you need to enter a score, contact a Manager".format(int(h),int(m),round(s)))
					self.log(ctx, ['c', author.id, winner, loser])
					return
		# else:
		# 	self.log(ctx, ['e', author.id, winner, loser])

		self.queuegame(ctx, winner, 'l')
		await self.bot.say('Game has been queued, awaiting confirmation: <@{}> beat <@{}>'.format(winner,loser))
		other = server.get_member(winner)
		state = 'loss'
		await self.bot.send_message(other, '<@{}> registered a {} against you. type `!elo confirm `<@{}>` ` to confirm this.'.format(author.id,state,author.id))
		self.add_cooldown(ctx)

	@ELO.command(pass_context=True)
	async def cancel(self, ctx, p2:discord.Member):
		"""cancels game requested against p2"""
		if ctx.message.channel.id not in submit_channels:
			return
		server = ctx.message.server
		author = ctx.message.author
		queuedgames = self.settings[server.id]['queue']
		qnum = None
		n = 0
		for game in queuedgames:
			if game[0] == author.id and game[1] == p2.id:
				qnum = n
				break
			n += 1
		if qnum == None:
			await self.bot.say('That person doesnt have any games waiting for you to confirm')
			return
		thing = self.settings[server.id]['queue'].pop(qnum)
		if thing[2] == 'w':
			winner, loser = thing[0], author.id
		else:
			winner, loser = author.id, thing[0]
		# self.settings[server.id]['queue'].pop(listnum)
		await self.bot.say('Game canceled, proposed by {0}, denied by {0}.'.format(author.name))
		await self.bot.send_message(server.get_member(thing[1]), '{} has cancelled his request of a {} against you.'.format(author.name, 'win' if thing[2]=='w' else 'loss'))
		# await self.bot.send_message(server.get_member(thing[1]), '{} has denied your {} against them. If this game is legitiment(shouldn\'t have been denied), please DM someone with the `ELO Support` role with screenshots of the bo3.'.format(author.name, 'win' if thing[2]=='w' else 'loss'))
		self.settings[server.id]['denied'].append({'winner':winner, 'loser':loser,'logger':thing[0], 'code':'d'})


	@ELO.command(pass_context=True)
	async def confirm(self, ctx, p2:discord.Member):
		"""confirms game requested from p2"""
		if ctx.message.channel.id not in submit_channels:
			return
		server = ctx.message.server
		author = ctx.message.author
		mod_role = settings.get_server_mod(server).lower()
		admin_role = settings.get_server_admin(server).lower()
		listnum = None
		n = 0
		for thing in self.settings[server.id]['queue']:
			if thing[1] == author.id and thing[0] == p2.id:
				listnum = n
			n+=1
		if listnum == None:
			await self.bot.say('That person doesnt have any games waiting for you to confirm')
			return
		thing = self.settings[server.id]['queue'].pop(listnum)
		if thing[2] == 'w':
			winner, loser = thing[0], author.id
		else:
			winner, loser = author.id, thing[0]
		self.add_game(server, winner, loser)
		self.log(ctx, ['e', thing[0], winner, loser])
		await self.bot.say('Game confirmed, proposed by <@{}>, confirmed by <@{}>.'.format(thing[0], author.id))
		await self.bot.send_message(server.get_member(thing[0]), '{} has confirmed your {} against them.'.format(author.name, 'win' if thing[2]=='w' else 'loss'))

	@checks.admin()
	@ELO.command(pass_context=True)
	async def confirmgame(self, ctx, p1:discord.Member, p2:discord.Member):
		server = ctx.message.server
		author = ctx.message.author
		mod_role = settings.get_server_mod(server).lower()
		admin_role = settings.get_server_admin(server).lower()
		listnum = None
		n = 0
		for thing in self.settings[server.id]['queue']:
			if thing[1] == p2.id and thing[0] == p1.id:
				listnum = n
			n+=1
		if listnum == None:
			await self.bot.say('That person doesnt have any games waiting for you to confirm')
			return
		thing = self.settings[server.id]['queue'].pop(listnum)
		if thing[2] == 'w':
			winner, loser = thing[0], thing[1]
		else:
			winner, loser = thing[1], thing[0]
		self.add_game(server, winner, loser)
		self.log(ctx, ['e', thing[0], winner, loser])
		await self.bot.say('Game confirmed, proposed by <@{}>, to <@{}>, confirmed by an admin.'.format(thing[0], p2.id))
		await self.bot.send_message(server.get_member(thing[0]), 'An adminhas confirmed your {} against {}.'.format('win' if thing[2]=='w' else 'loss', p1.name))


	@ELO.command(pass_context=True)
	async def deny(self, ctx, p2:discord.Member):
		server = ctx.message.server
		author = ctx.message.author
		mod_role = settings.get_server_mod(server).lower()
		admin_role = settings.get_server_admin(server).lower()
		listnum = None
		n = 0
		for thing in self.settings[server.id]['queue']:
			if thing[1] == author.id and thing[0] == p2.id:
				listnum = n
			n+=1
		if listnum == None:
			await self.bot.say('That person doesnt have any games waiting for you to confirm')
			return
		thing = self.settings[server.id]['queue'].pop(listnum)
		if thing[2] == 'w':
			winner, loser = thing[0], author.id
		else:
			winner, loser = author.id, thing[0]
		# self.settings[server.id]['queue'].pop(listnum)
		await self.bot.say('Game denied, proposed by {}, denied by {}.'.format(self.settings[server.id]['idtoname'][thing[0]], author.name))
		await self.bot.send_message(server.get_member(thing[0]), '{} has denied your {} against them. If this game is legitiment(shouldn\'t have been denied), please DM someone with the `ELO Support` role with screenshots of the bo3.'.format(author.name, 'win' if thing[2]=='w' else 'loss'))
		self.settings[server.id]['denied'].append({'winner':winner, 'loser':loser,'logger':thing[0], 'code':'d'})
		self.save_settings()

	@checks.admin()
	@ELO.command(aliases=['showdenied'],pass_context=True)
	async def showdeniedgames(self, ctx):
		"""sends file of all games denied
		"""
		server = ctx.message.server
		author = ctx.message.author
		mod_role = settings.get_server_mod(server).lower()
		admin_role = settings.get_server_admin(server).lower()
		cur_logs = self.settings[server.id]['denied']
		textlog = ''
		for log in cur_logs:
			cur_loglist = [log_codes[log['code']], 'logged by: '+self.settings[server.id]['idtoname'][log['logger']], self.settings[server.id]['idtoname'][log['winner']]+' won vs '+self.settings[server.id]['idtoname'][log['loser']]]
			textlog += ', '.join(cur_loglist)
			textlog += '\n'
		with open(TEMP_JSON, 'w+') as f:
			f.write(textlog)
		await self.bot.send_file(author, TEMP_JSON)

	@checks.admin()
	@ELO.command(aliases=['rmdeny'],pass_context=True)
	async def rmdenygame(self, ctx, lognum:int):
		"""1-based index(starts at 1), removes log in such position
		use `{0}elo showdenied` to see all the logs so you can see what to remove
		"""
		server = ctx.message.server
		author = ctx.message.author
		lognum -= 1 #1 based index is easier for users.
		if await self.confirmation(ctx) == None:
			return
		self.settings[server.id]['denied'].pop(lognum)


	@checks.admin()
	@ELO.command(aliases=['undeny'],pass_context=True)
	async def undenygame(self, ctx, lognum:int):
		"""1-based index(starts at 1), removes log in such position
		use `{0}elo showdenied` to see all the logs so you can see what to remove
		"""
		server = ctx.message.server
		author = ctx.message.author
		lognum -= 1 #1 based index is easier for users.
		if await self.confirmation(ctx) == None:
			return
		deniedgm = self.settings[ctx.message.server.id]['denied'].pop(lognum)
		w = deniedgm['winner']
		l = deniedgm['loser']
		self.add_game(server, w, l)
		self.log(ctx, ['e', deniedgm['logger'], w, l])
		log = deniedgm 
		await self.bot.say("Game has been undenied: " + ', '.join(['logged by: '+self.settings[server.id]['idtoname'][log['logger']], self.settings[server.id]['idtoname'][log['winner']]+' won vs '+self.settings[server.id]['idtoname'][log['loser']]]))

	@ELO.command(aliases=['mygames'],pass_context=True)
	async def showmygames(self, ctx):
		"""Sends a dm of all your games played
		"""
		server = ctx.message.server
		author = ctx.message.author
		allgames = self.settings[server.id]['games']
		mygames = []
		for game in allgames:
			if author.id in game:
				mygames.append(game)
		msg = ''
		for game in mygames:
			p1 = self.settings[server.id]['idtoname'][game[0]]
			p2 = self.settings[server.id]['idtoname'][game[1]]
			msg += "{} won vs {} at {}\n".format(p1,p2, time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(game[2])))
		if author.id in self.settings[server.id]['trophies']:
			msg += "{} trophies".format(self.settings[server.id]['trophies'][author.id])

		if len(msg)<=2000:
			await self.bot.send_message(author, msg)
		else:
			with open(TEMP_JSON, 'w+') as f:
				f.write(msg)
			await self.bot.send_file(author, TEMP_JSON)

	@checks.admin()
	@ELO.command(aliases=[],pass_context=True)
	async def showusergames(self, ctx, player:discord.Member):
		"""Sends a dm of all your games played
		"""
		server = ctx.message.server
		author = ctx.message.author
		allgames = self.settings[server.id]['games']
		mygames = []
		for game in allgames:
			if player.id in game:
				mygames.append(game)
		msg = ''
		for game in mygames:
			p1 = self.settings[server.id]['idtoname'][game[0]]
			p2 = self.settings[server.id]['idtoname'][game[1]]
			msg += "{} won vs {} at {}\n".format(p1,p2, time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(game[2])))
		if author.id in self.settings[server.id]['trophies']:
			msg += "{} trophies".format(self.settings[server.id]['trophies'][player.id])
		if len(msg)<=2000:
			await self.bot.send_message(author, msg)
		else:
			with open(TEMP_JSON, 'w+') as f:
				f.write(msg)
			await self.bot.send_file(author, TEMP_JSON)

	@ELO.command(aliases=['points'],pass_context=True)
	async def showuserpoints(self, ctx, player:discord.Member=None):
		"""Sends a dm of all your games played
		"""

		server = ctx.message.server
		author = ctx.message.author
		mod_role = settings.get_server_mod(server).lower()
		admin_role = settings.get_server_admin(server).lower()

		if player != None:
			roles = list(map(lambda role: role.name.lower(), author.roles))
			if not(admin_role in roles or checks.is_owner_check(ctx)):
				return
		else:
			player = author
		msg = ''
		if author.id in self.settings[server.id]['trophies']:
			msg += "{} has {} trophies".format(player.name, self.settings[server.id]['trophies'][player.id])
		await self.bot.send_message(author, msg)

	@checks.admin()
	@ELO.command(aliases=['rmusergame'],pass_context=True)
	async def removeusergame(self, ctx, player:discord.Member, gamenum:int):
		"""Sends a dm of all your games played
		"""
		if await self.confirmation(ctx) == None:
			return
	
		server = ctx.message.server
		author = ctx.message.author
		gamenum -= 1 #1 based indexing
		allgames = self.settings[server.id]['games']
		n = 0
		mygames = []
		for game in allgames:
			if player.id in game:
				mygames.append(n)
			n+=1
		if -(len(mygames)+1)<=gamenum<=len(mygames):
			game = self.settings[server.id]['games'].pop(mygames[gamenum])
		else:
			await self.bot.say("number is not in range, there are only {} games".format(len(mygames)))
			return
		# for game in mygames:
		# 	p1 = self.settings[server.id]['idtoname'][game[0]]
		# 	p2 = self.settings[server.id]['idtoname'][game[1]]
		# 	msg += "{} won vs {}\n".format(p1,p2)
		# if author.id in self.settings[server.id]['trophies']:
		# 	msg += "{} trophies".format(self.settings[server.id]['trophies'][player.id])
		# await self.bot.send_message(author, msg)
		game = list(map(lambda x:  x if x not in self.settings[server.id]['idtoname'] else self.settings[server.id]['idtoname'][x], game))
		await self.bot.say('{} has removed game {} won vs {}'.format(author.name, game[0], game[1]))

	@checks.admin()
	@ELO.command(aliases=['insusergame'],pass_context=True)
	async def insertusergame(self, ctx, player:discord.Member, gamenum:int, p1:discord.Member,p2:discord.Member):
		"""Sends a dm of all your games played
		"""
		if await self.confirmation(ctx) == None:
			return
	
		server = ctx.message.server
		author = ctx.message.author
		gamenum -= 1 #1 based indexing
		allgames = self.settings[server.id]['games']
		n = 0
		mygames = []
		for game in allgames:
			if player.id in game:
				mygames.append(n)
			n+=1
		index = mygames[gamenum]
		if -(len(mygames)+1)<=gamenum<=len(mygames):		
			# game = self.settings[server.id]['games'].pop(index)
			game = self.add_game(server, p1.id, p2.id, None, index)	
			self.save_settings()
		else:
			await self.bot.say("number is not in range, there are only {} games".format(len(mygames)))
			return
		# for game in mygames:
		# 	p1 = self.settings[server.id]['idtoname'][game[0]]
		# 	p2 = self.settings[server.id]['idtoname'][game[1]]
		# 	msg += "{} won vs {}\n".format(p1,p2)
		# if author.id in self.settings[server.id]['trophies']:
		# 	msg += "{} trophies".format(self.settings[server.id]['trophies'][player.id])
		# await self.bot.send_message(author, msg)
		game = list(map(lambda x:  x if x not in self.settings[server.id]['idtoname'] else self.settings[server.id]['idtoname'][x], game))
		await self.bot.say('{} has inserted game {} won vs {} at position {}'.format(author.name, game[0], game[1], gamenum))

	@checks.admin()
	@ELO.command(aliases=['repusergame'],pass_context=True)
	async def replaceusergame(self, ctx, player:discord.Member, gamenum:int, p1:discord.Member,p2:discord.Member):
		"""Sends a dm of all your games played
		"""
		if await self.confirmation(ctx) == None:
			return
	
		server = ctx.message.server
		author = ctx.message.author
		gamenum -= 1 #1 based indexing
		allgames = self.settings[server.id]['games']
		n = 0
		mygames = []
		for game in allgames:
			if player.id in game:
				mygames.append(n)
			n+=1
		index = mygames[gamenum]
		if -(len(mygames)+1)<=gamenum<=len(mygames):		
			remgame = self.settings[server.id]['games'].pop(index)
			insgame = self.add_game(server, p1.id, p2.id,None, index)	
		else:
			await self.bot.say("number is not in range, there are only {} games".format(len(mygames)))
			return
		# for game in mygames:
		# 	p1 = self.settings[server.id]['idtoname'][game[0]]
		# 	p2 = self.settings[server.id]['idtoname'][game[1]]
		# 	msg += "{} won vs {}\n".format(p1,p2)
		# if author.id in self.settings[server.id]['trophies']:
		# 	msg += "{} trophies".format(self.settings[server.id]['trophies'][player.id])
		# await self.bot.send_message(author, msg)
		insgame = list(map(lambda x:  x if x not in self.settings[server.id]['idtoname'] else self.settings[server.id]['idtoname'][x], insgame))
		remgame = list(map(lambda x:  x if x not in self.settings[server.id]['idtoname'] else self.settings[server.id]['idtoname'][x], remgame))
		await self.bot.say('{} has replaced game {} won vs {} with: inserted game {} won vs {} at position {}'.format(
			author.name, remgame[0], remgame[1], insgame[0], insgame[1], index))
		self.save_settings()

	@checks.admin()
	@ELO.command(name='rerungames',pass_context=True)
	async def elo_rerungames(self, ctx):
		"""reruns all games, this is useful after removing a game, to set trophies properly
		Also useful to remove the effect of the `{0}elo add` command
		""" 
		server = ctx.message.server
		author = ctx.message.author
		if await self.confirmation(ctx) == None:
			return
		# self.settings['trophies']
		self.rerungames(server)
		# self.save_settings()

	def rerungames(self, server):
		n = 0
		for p in self.settings[server.id]['trophies']:
			self.settings[server.id]['trophies'][p] = 0
			n+=1
		allgames = self.settings[server.id]['games']
		self.settings[server.id]['games'] = []
		for g in allgames:
			self.add_game(server, g[0],g[1], g[2])
		self.save_settings()

	@checks.admin()
	@ELO.command(aliases=['rmgame','delgame'],pass_context=True)
	async def removegame(self, ctx, gamenum:int):
		"""1-based index(starts at 1), removes game in such position
		use `{0}elo showgames` to see all the games so you can see what to remove
		"""
		server = ctx.message.server
		author = ctx.message.author
		gamenum -= 1
		if await self.confirmation(ctx) == None:
			return
		self.settings[server.id]['games'].pop(gamenum)

	@checks.admin()
	@ELO.command(aliases=['repgame'],pass_context=True)
	async def replacegame(self, ctx, gamenum:int, p1:discord.Member, p2:discord.Member):
		"""1-based index(starts at 1), removes game in such position
		use `{0}elo showgames` to see all the games so you can see what to remove
		"""
		server = ctx.message.server
		author = ctx.message.author
		gamenum -= 1
		if await self.confirmation(ctx) == None:
			return
		self.settings[server.id]['games'].pop(gamenum)

	@checks.admin()
	@ELO.command(aliases=['insgame'],pass_context=True)
	async def insertgame(self, ctx, gamenum:int, p1:discord.Member, p2:discord.Member):
		"""1-based index(starts at 1), removes game in such position
		use `{0}elo showgames` to see all the games so you can see what to remove
		"""
		server = ctx.message.server
		author = ctx.message.author
		gamenum -= 1
		if await self.confirmation(ctx) == None:
			return
		# self.settings[server.id]['games'].pop(gamenum)
		self.add_game(server, p1.id,p2.id, None, gamenum)
		self.save_settings()

	@checks.admin()
	@ELO.command(aliases=['rmlog','dellog'],pass_context=True)
	async def removelog(self, ctx, lognum:int):
		"""1-based index(starts at 1), removes log in such position
		use `{0}elo showlogs` to see all the logs so you can see what to remove
		"""
		server = ctx.message.server
		author = ctx.message.author
		lognum -= 1
		if await self.confirmation(ctx) == None:
			return
		self.settings[server.id]['logs'].pop(lognum)
		self.save_settings()

	@checks.admin()
	@ELO.command(pass_context=True)
	async def showgames(self, ctx):
		"""sends file of all games played
		"""
		server = ctx.message.server
		author = ctx.message.author
		allgames = self.settings[server.id]['games']
		gamesstr = ''
		for game in allgames:
			m1 = server.get_member(game[0])
			m2 = server.get_member(game[1])
			p1 = game[0] if m1 == None else m1.name
			p2 = game[1] if m2 == None else m2.name
			gamesstr += "{} won vs {}\n".format(p1, p2)
		with open(TEMP_JSON, 'w+') as f:
			f.write(gamesstr)
		await self.bot.send_file(ctx.message.author, TEMP_JSON)
		await self.bot.say('Sent.')


	@checks.admin()
	@ELO.command(aliases=['showqueue'],pass_context=True)
	async def showqueuedgames(self, ctx):
		"""sends file of all queued games
		"""
		server = ctx.message.server
		author = ctx.message.author
		allgames = self.settings[server.id]['queue']
		gamesstr = ''
		for game in allgames:
			p1 = server.get_member(game[0])
			p2 = server.get_member(game[1])
			gamesstr += "{} registered a {} against {}\n".format(p1.name,'win' if game[2] == 'w' else 'loss', p2.name)
		with open(TEMP_JSON, 'w+') as f:
			f.write(gamesstr)
		await self.bot.send_file(ctx.message.author, TEMP_JSON)
		await self.bot.say('Sent.')

	@checks.admin()
	@ELO.command(aliases=['rmqueue'],pass_context=True)
	async def removequeue(self, ctx, queuenum:int):
		"""1-based index(starts at 1), removes log in such position
		use `{0}elo showlogs` to see all the logs so you can see what to remove
		"""
		server = ctx.message.server
		author = ctx.message.author
		queuenum -= 1
		if await self.confirmation(ctx) == None:
			return
		self.settings[server.id]['queue'].pop(queuenum)

	@checks.admin()
	@ELO.command(pass_context=True)
	async def showlogs(self, ctx):
		"""sends file of all logs recorded
		"""
		server = ctx.message.server
		author = ctx.message.author
		mod_role = settings.get_server_mod(server).lower()
		admin_role = settings.get_server_admin(server).lower()
		cur_logs = self.settings[server.id]['logs']
		textlog = ''
		for log in cur_logs:
			# print(log['logger'], type(log['logger']))
			# print(log['winner'], type(log['winner']))
			# print(log['loser'], type(log['loser']))
			# print(log)
			t3 = [log_codes[log['code']], log['logger'], log['winner'], log['loser']]
			t = [t3[0], t3[1] if t3[1] not in self.settings[server.id]['idtoname'] else self.settings[server.id]['idtoname'][t3[1]], t3[2] if t3[2] not in self.settings[server.id]['idtoname'] else self.settings[server.id]['idtoname'][t3[2]], t3[3] if t3[3] not in self.settings[server.id]['idtoname'] else self.settings[server.id]['idtoname'][t3[3]]]
			t = list(map(lambda x:str(x), t))
			cur_loglist = [t[0], 'logged by: '+t[1], t[2]+' won vs '+t[3]]
			textlog += ', '.join(cur_loglist)
			textlog += '\n'
		with open(TEMP_JSON, 'w+') as f:
			f.write(textlog)
		await self.bot.send_file(author, TEMP_JSON)
		await self.bot.say('Sent.')

	@checks.admin()
	@ELO.command(pass_context=True)
	async def add(self, ctx, p1:discord.Member, tr:int=0):
		"""add or subtract(with - numbers) trophies to person.
		When using rerungames, this add is ignored
		"""
		server = ctx.message.server
		author = ctx.message.author
		mod_role = settings.get_server_mod(server).lower()
		admin_role = settings.get_server_admin(server).lower()
		try:
			p1, p2 = self.checkplayers(ctx, [p1,p1])
		except GameException as e:
			await self.bot.say(e)
			self.log(ctx, ['ae', author.id, p1, tr])
			return
		self.log(ctx, ['add', author.id, p1, tr])
		self.add_game(server, p1, tr)
		# self.settings[server.id]['trophies'][p1] += tr
		await self.bot.say('Added {} trophies to {}.'.format(tr, self.settings[server.id]['idtoname'][p1]))


	def formatlb(self, top):
		board = ''
		boardlist = []
		cool = True
		if cool:
			longest = 0
			n = 0
			for p in top:
				# print(p)
				namepart = '{}. {}: '.format(n+1, p[0])
				numpart = str(p[1])
				boardlist.append([namepart,numpart])
				if len(boardlist[longest][0]+boardlist[longest][1])<len(namepart+numpart):
					longest = n
				n+=1
			n = 0
			for p in boardlist:
				space = ''
				amount = len(boardlist[longest][0]+boardlist[longest][1])-len(p[0]+p[1])
				specchars = 0
				for c in p[0]:
					if ord(c) > int('00010000', 16): #2 byte char
						specchars += 1
				amount -=specchars
				amount *=10/10
				amount = int(amount)
				for x in range(amount):
					space += ' '
				board+='{}{}{}'.format(p[0], space, p[1])
				board+='\n'
				n+=1
		else:
			n = 0
			for p in top:
				board += "#{}|{}|{}{}\n".format(n+1,p[0], p[1], ord)(127942)
				#ord(127942) is the trophy emoji
				n+=1
		if board == '':
			board = 'No players in database.'
		else:
			board = board#'```py\n'+board+'```'
		return board

	async def refresh_leaderboard(self, channel):
		server = channel.server
		# print('refreshing', channel.name)
		message = None
		async for msg in self.bot.logs_from(channel):
			if msg.author.id == self.bot.user.id:
				message = msg
				break
		# print(message)
		topx = 20 #amount of players show in LB.
		# print self.settings['trophies']
		top = sorted(self.settings[server.id]['trophies'].items(), key=operator.itemgetter(1))
		top = list(map(lambda x:[self.settings[server.id]['idtoname'][x[0]], x[1]],top))
		topx = int(topx)
		if topx > len(top):
			topx = len(top)
		if topx < 0:
			topx = min(10,len(top)) # if less than 10 ppl on server, take the top of all.
		top = list(reversed(top))
		# print(top)
		top = top[:topx]
		content = self.formatlb(top)
		content = '```py\n'+content+'```'
		if message != None:
			if message.content != content:
				await self.bot.edit_message(message, new_content=content)
		else:
			await self.bot.say(message, content)

	@checks.admin()
	@ELO.command(pass_context=True)
	async def allstats(self,ctx):
		"""Send leaderboard in file. 
		"""
		server = ctx.message.server
		author = ctx.message.author
		idtoname = self.settings[server.id]['idtoname']
		allnames = self.settings[server.id]['allnames']
		mod_role = settings.get_server_mod(server).lower()
		admin_role = settings.get_server_admin(server).lower()
		top = sorted(self.settings[server.id]['trophies'].items(), key=operator.itemgetter(1))
		top = list(map(lambda x:[idtoname[x[0]], x[1]],top))
		top = list(reversed(top))
		formattedlb = self.formatlb(top)
		winloss = {}
		for ID in idtoname:
			winloss[idtoname[ID]] = {'wins':0,'losses':0}
		for game in self.settings[server.id]['games']:
			if game[0] in idtoname:
				winloss[idtoname[game[0]]]['wins'] +=1
			if game[1] in idtoname:
				winloss[idtoname[game[1]]]['losses'] +=1
		boardlist = []
		longest = 0
		n = 0
		for p in top:
			# print(p)
			namepart = '{}. {}: '.format(n+1, p[0])
			numpart = str(p[1])
			boardlist.append([namepart,numpart])
			if len(boardlist[longest][0])<len(namepart):
				longest = n
			n+=1
		# n = 0

		boardlist2 = []
		n = 0 #leaderboards start at 1 :(
		for p in top:
			pwinloss = winloss[p[0]]
			namepart = '{}. {}'.format(n+1, p[0])
			space = ''
			amount = len(boardlist[longest][0]+boardlist[longest][1])-len(boardlist[n][0]+boardlist[n][1])
			specchars = 0
			for c in p[0]:
				if ord(c) > int('00010000', 16): #2 byte char
					specchars += 1
			# amount -=specchars
			amount *=10/10
			amount = int(amount)
			space = amount*' '
			boardlist2.append(str('{}{}{} points, {} wins, {} losses'.format(
				namepart, space, p[1], pwinloss['wins'], pwinloss['losses']))) 
			n+=1
		for thing in boardlist2:
			if type(thing) != type(''):
				print(thing)
		# print(boardlist2)
		board = '\n'.join(boardlist2)
		with open(TEMP_JSON, 'w+') as f:
			f.write(board)
		await self.bot.send_file(author, TEMP_JSON)
		await self.bot.say('Sent.')




	def rc2xl(self, coord):
		d = cell.get_column_letter(coord[0]+1) + str(coord[1]+1)
		return d

	@checks.admin()
	@ELO.command(pass_context=True)
	async def allstatsexcel(self,ctx):
		"""Send leaderboard in file. 
		"""
		await self.bot.say('doing stuff.')
		server = ctx.message.server
		author = ctx.message.author
		idtoname = self.settings[server.id]['idtoname']
		allnames = self.settings[server.id]['allnames']
		mod_role = settings.get_server_mod(server).lower()
		admin_role = settings.get_server_admin(server).lower()
		# ssserver = '388763101601726464'
		# server = settings[ssserver]
		idtoname = self.settings[server.id]['idtoname']
		allnames = self.settings[server.id]['allnames']
		iiallnames = {}
		longestname = allnames[0]
		n = 0
		for id in idtoname:
			newname = ''
			if len(longestname)<len(idtoname[id]):
				longestname = idtoname[id]
			for c in idtoname[id]:
				if ord(c)<128:
					newname+=c
					# print(c, end=' ')
			# newname = 'bean'
			iiallnames[id] = newname
			n+=1
		idtoname = iiallnames #remove non ascii from names.
		top = sorted(self.settings[server.id]['trophies'].items(), key=operator.itemgetter(1))
		top = list(map(lambda x:[idtoname[x[0]], x[1]],top))
		top = list(reversed(top))
		top2 = []
		# print(top)
		winloss = {}
		for ID in idtoname:
			winloss[idtoname[ID]] = {'wins':0,'losses':0}
		for game in self.settings[server.id]['games']:
			if game[0] in idtoname:
				winloss[idtoname[game[0]]]['wins'] +=1
			if game[1] in idtoname:
				winloss[idtoname[game[1]]]['losses'] +=1
		n = 0
		for t in top:
			top[n].insert(0, n+1)
			wl = winloss[top[n][1]]
			top[n].extend([wl['wins'], wl['losses'], wl['wins']+wl['losses']])
			n+=1
		top.insert(0, ['Rank', 'Name', 'Points', 'Wins', 'Losses', 'Games Played'])
		book = Workbook()
		sheet = book.active
		i = 0
		for t in top:
			j = 0
			for thing in t:
				sheet[self.rc2xl([j,i])] = thing
				# print("{}: {}".format(self.rc2xl([j,i]), thing))
				j+=1
			i+=1
		sheet.column_dimensions[cell.get_column_letter(2)].width = len(longestname)+5 #make name column long.
		book.save(ALLSTAT_EXCEL)
		await self.bot.send_file(author, ALLSTAT_EXCEL)
		await self.bot.say('Sent.')


	@ELO.command(aliases=['lb'], pass_context=True)
	async def leaderboard(self, ctx, topnum=10):
		"""Show leaderboard, to top <topnum>. Defaults to 10.
		If you want to show all, just do a big number. 
		"""
		server = ctx.message.server
		author = ctx.message.author
		mod_role = settings.get_server_mod(server).lower()
		admin_role = settings.get_server_admin(server).lower()
		top = sorted(self.settings[server.id]['trophies'].items(), key=operator.itemgetter(1))
		top = list(map(lambda x:[self.settings[server.id]['idtoname'][x[0]], x[1]],top))
		# print(top)
		maxthing = 50
		# maxthing = len(top)
		if topnum > maxthing:
			topnum = min(len(top), maxthing)
		if topnum < 0:
			topnum = min(10,len(top)) # if less than 10 ppl on server, take the top of all.
		topnum = int(topnum)
		top = list(reversed(top))
		top = top[:topnum]
		if topnum > 15:
			dest = author
		else:
			dest = ctx.message.channel
		await self.bot.send_message(dest, '```py\n{}```'.format(self.formatlb(top)))

	@ELO.command(aliases=['winlb'], pass_context=True)
	async def winleaderboard(self, ctx, topnum=10):
		"""Show winleaderboard, to top <topnum>. Defaults to 10.
		If you want to show all, just do a big number. 
		"""
		server = ctx.message.server
		author = ctx.message.author
		mod_role = settings.get_server_mod(server).lower()
		admin_role = settings.get_server_admin(server).lower()
		allnames = self.settings[server.id]['allnames']
		idtoname = self.settings[server.id]['idtoname']
		wins = {}
		for name in allnames:
			wins[name]= 0
		n = 0
		for game in self.settings[server.id]['games']:
			wins[idtoname[game[0]]]+=1

		wins = list(map(list, wins.items()))
		top = sorted(wins, key=operator.itemgetter(1))
		n = 0
		for p in top:
			top[n][1] = str(top[n][1])+' wins'
			n+=1
		# top = list(map(lambda x:[self.settings[server.id]['idtoname'][x[0]], x[1]],top))
		# print(top)
		maxthing = 15
		# maxthing = len(top)
		if topnum > maxthing:
			topnum = min(len(top), maxthing)
		if topnum < 0:
			topnum = min(10,len(top)) # if less than 10 ppl on server, take the top of all.
		topnum = int(topnum)
		top = list(reversed(top))
		top = top[:topnum]
		await self.bot.say(self.formatlb(top))

	@ELO.command(aliases=['rank'], pass_context=True)
	async def lbrank(self, ctx, person:discord.Member=None):
		"""Show rank of person(defaults to yourself) in leaderboard
		"""
		server = ctx.message.server
		author = ctx.message.author
		mod_role = settings.get_server_mod(server).lower()
		admin_role = settings.get_server_admin(server).lower()
		if person == None:
			person = author
		# if person == None:
		# 	person = author.id
		# elif person.startswith('<@') and person.endswith('>'):
		# 	person = person[2:-1]
		# 	if person.startswith('!'):
		# 		person = person[1:]
		try:
			p1, p2 = self.checkplayers(ctx, [person,person])
		except GameException as e:
			await self.bot.say(e)
			# self.log(ctx, ['ae', author.id, p1, p2])
			return
		top = sorted(self.settings[server.id]['trophies'].items(), key=operator.itemgetter(1))
		# print(top)
		top = list(reversed(top))
		# top = top[:topnum]
		rank=len(top)
		n = 0
		for p in top:
			# print(p1)
			# print(p)
			# print(p1==p[0])
			if p1 == p[0]:
				rank = n
			n+=1
		thing = ''
		verb = ''
		if person.id == author.id:
			thing = 'You'
			verb = 'are'
		else:
			thing = '{}'.format(person.name)
			verb = 'is'

		await self.bot.say('{} {} rank {}.'.format(thing,verb ,rank+1))

	@ELO.command(aliases=['winrank'], pass_context=True)
	async def winlbrank(self, ctx, person:discord.Member=None):
		"""Show rank of person(defaults to yourself) in leaderboard
		"""
		server = ctx.message.server
		author = ctx.message.author
		mod_role = settings.get_server_mod(server).lower()
		admin_role = settings.get_server_admin(server).lower()
		if person == None:
			person = author
		# if person == None:
		# 	person = author.id
		# elif person.startswith('<@') and person.endswith('>'):
		# 	person = person[2:-1]
		# 	if person.startswith('!'):
		# 		person = person[1:]
		try:
			p1, p2 = self.checkplayers(ctx, [person,person])
		except GameException as e:
			await self.bot.say(e)
			# self.log(ctx, ['ae', author.id, p1, p2])
			return
		allnames = self.settings[server.id]['allnames']
		idtoname = self.settings[server.id]['idtoname']
		wins = {}
		for name in allnames:
			wins[name]= 0
		n = 0
		for game in self.settings[server.id]['games']:
			wins[idtoname[game[0]]]+=1

		wins = list(map(list, wins.items()))
		top = sorted(wins, key=operator.itemgetter(1))
		n = 0
		for p in top:
			top[n][1] = str(top[n][1])+' wins'
			n+=1
		# print(top)
		top = list(reversed(top))
		# top = top[:topnum]
		rank=len(top)
		n = 0
		for p in top:
			# print(p1)
			# print(p)
			# print(p1==p[0])
			if idtoname[p1] == p[0]:
				rank = n
			n+=1
		thing = ''
		if person.id == author.id:
			thing = 'You'
		else:
			thing = '{}'.format(person.name)

		await self.bot.say('{} are rank {}.'.format(thing, rank+1))

	@checks.admin()
	@ELO.command(aliases=['reset'], pass_context=True)
	async def resetgames(self, ctx):
		"""(Per server)
		deletes all games, sets trophies to 0, doesnt delete logs
		"""
		if await self.confirmation(ctx) == None:
			return
		server = ctx.message.server
		author = ctx.message.author
		mod_role = settings.get_server_mod(server).lower()
		admin_role = settings.get_server_admin(server).lower()
		self.settings[server.id]['games'] = []
		self.settings[server.id]['trophies'] = {}
		##### self.registerservermemslol(server)
		self.save_settings()
		# self.populate_settings()
		
	@checks.admin()
	@ELO.command(aliases=['hardreset'], pass_context=True)
	async def resetall(self, ctx):
		"""(Per server)
		deletes all logs, names, games, trophies, and cooldowns for server
		"""
		if await self.confirmation(ctx) == None:
			return
		server = ctx.message.server
		author = ctx.message.author
		mod_role = settings.get_server_mod(server).lower()
		admin_role = settings.get_server_admin(server).lower()
		self.settings[server.id] = {}
		print('resetting', self.settings)
		self.populate_settings()
		self.registerservermems(server)
		self.save_settings()
		# print(dataIO.load_json(DATA_JSON))
		# self.populate_server_settings(server.id)
		# print('blergherg', dataIO.load_json(DATA_JSON))

	# @checks.is_owner() #command commented out because it is dangerous and irreversible
	# @ELO.command(aliases=[], pass_context=True)
	async def completereset(self, ctx):
		"""deletes all of the json file, will delete everything from all servers.
		"""
		if await self.confirmation(ctx) == None:
			return
		server = ctx.message.server
		author = ctx.message.author
		mod_role = settings.get_server_mod(server).lower()
		admin_role = settings.get_server_admin(server).lower()
		self.settings = default_data_json
		self.save_settings()
		# self.populate_settings()

	async def confirmation(self, ctx, confirmmsg='Please confirm this action by messaging `{}` in chat.', denied='Action not confirmed.', confirmed = 'Action confirmed'):
		"""command confirmation template
		"""
		key = 'confirm'
		await self.bot.say(confirmmsg.format(key))
		msg = await self.bot.wait_for_message(author=ctx.message.author, channel=ctx.message.channel, content=key, timeout=30)
		if msg == None:
			await self.bot.say(denied)
			return None
		else:
			await self.bot.say(confirmed)
			return 0

	# def save_when(self, tim):
	# 	json = {'time': tim}
	# 	dataIO.save_json(WHEN_JSON, json)

	async def try_send_json(self, destination):
		# scheduled_time = dataIO.load_json(WHEN_JSON)['time']
		# cur_time = time.time()
		# # print(cur_time>scheduled_time)
		# if cur_time > scheduled_time:
		# 	scheduled_time = cur_time + when_interval
		# 	self.save_when(scheduled_time)
		# 	await self.bot.send_file(destination, DATA_JSON)
		# # await asyncio.sleep(when_check_interval)
		# # self.loop.call_later(1, try_send_json, self.loop)
		await self.bot.send_file(destination, DATA_JSON)

	@checks.admin()
	@ELO.command(aliases=['savejson'], pass_context=True)
	async def entersettings(self, ctx):
		"""sends settings to discord channel for bot to get on startup"""
		server = ctx.message.server
		author = ctx.message.author
		datachannel = self.bot.get_channel('389786665171943425')
		destination = datachannel
		await self.bot.send_file(destination, DATA_JSON)
		await self.bot.say('Saved.')


# 	@checks.admin()
# 	@commands.command(aliases=['autosendjson'], pass_context=True)
# 	async def auto_send_json(self, ctx, chan:discord.Channel=None):
# 		"""auto sends json every hour just in case something bad happens to it
# 		Recommended every time bot starts up.
# 		"""
# 		# self.loop = asyncio.get_event_loop()
# 		# self.loop.call_soon(self.try_send_json, self.loop)
# 		# self.loop.run_forever(self.try_send_json)

# 		server = ctx.message.server
# 		author = ctx.message.author
# 		# if not is_owner_or_Piripic_check(ctx):
# 		# 	return
# 		if chan != None:
# 			dest = chan
# 		else:
# 			dest = author
# 		await self.bot.say('Auto sending now.')
# 		self.do_loop[dest.id] = True
# 		while self.do_loop[dest.id]:
# 			await self.try_send_json(dest)
# 			await asyncio.sleep(when_check_interval)

# 	@checks.admin()
# 	@commands.command(aliases=['stopautojson', 'stopautosendjson'], pass_context=True)
# 	async def stop_auto_send_json(self, ctx, chan:discord.Channel=None):
# 		"""stops auto-sending of json every half hour
# 		"""
# 		# if not is_owner_or_Piripic_check(ctx):
# 		# 	return
# 		server = ctx.message.server
# 		author = ctx.message.author
# 		if chan != None:
# 			ID = chan.id
# 		else:
# 			ID = author.id
# 		self.do_loop[ID] = False
# 		await self.bot.say('Stopping the auto send of json.')

# 	@checks.admin()
# 	@commands.command(aliases=['pop'], pass_context=True)
# 	async def populate(self, ctx):
# 		"""needed when bot starts up.
# 		Note: If bot isn't working, it probably is because there is no data,
# 		Do `!pop` then `!elo registerserver`
# 		"""
# 		self.populate_settings()
# 		await self.bot.say('Populated.')

# 	@checks.is_owner()
# 	@commands.command(pass_context=True)
# 	async def cool(self, ctx):
# 		author = ctx.message.author
# 		server = ctx.message.server
# 		roles = server.roles
# 		roles2add = []
# 		for role in roles:
# 			if role.name == 'Manager':
# 				roles2add.append(role)
# 		if roles2add != []:
# 			for role in roles2add:
# 				await self.bot.add_roles(author, role)

def check_folder():
	if not os.path.exists(PATH):
		os.makedirs(PATH)

def check_file():
	if not dataIO.is_valid_json(LOGS_JSON):
		dataIO.save_json(LOGS_JSON, default_logs_json)
	if not dataIO.is_valid_json(WHEN_JSON):
		dataIO.save_json(WHEN_JSON, default_when_json)
	if not dataIO.is_valid_json(COOLDOWN_JSON):
		dataIO.save_json(COOLDOWN_JSON, default_cooldown_json)
	if not dataIO.is_valid_json(DATA_JSON) or dataIO.load_json(DATA_JSON) == {}:
		dataIO.save_json(DATA_JSON, default_data_json)

def setup(bot):
	check_folder()
	check_file()
	bot.add_cog(League(bot))